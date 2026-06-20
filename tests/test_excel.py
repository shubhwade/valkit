from __future__ import annotations

import os
import tempfile
from openpyxl import load_workbook

from telmus.core.result import ScanResult, ValuationResult, HealthResult, GrowthResult, RedFlag
from telmus.exporters.excel import ExcelExporter


def test_excel_export_creates_valid_file() -> None:
    # Construct a complete mock ScanResult
    result = ScanResult(
        ticker="AAPL",
        company="Apple Inc.",
        exchange="NASDAQ",
        scan_duration_ms=120,
        valuation=ValuationResult(
            pe_ratio=28.5,
            pb_ratio=10.2,
            ev_ebitda=20.1,
            vs_sector="expensive",
            flag="expensive relative to sector"
        ),
        health=HealthResult(
            piotroski_f=7,
            altman_z=4.5,
            debt_to_equity=0.8,
            current_ratio=1.5,
            interest_coverage=10.0,
            flag=None
        ),
        growth=GrowthResult(
            revenue_cagr_3y=0.12,
            pat_cagr_3y=0.15,
            margin_trend="stable margins",
            fcf_yield=0.05,
            flag=None
        ),
        red_flags=[
            RedFlag(type="negative_fcf", value=2, severity="medium")
        ],
        highest_concern="medium",
        analyst_brief="Strong fundamentals, but moderate concerns over FCF."
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        dest_path = os.path.join(tmpdir, "test_report.xlsx")

        # Export
        exporter = ExcelExporter()
        exporter.export(result, dest_path)

        # Verify file exists
        assert os.path.exists(dest_path)

        # Load workbook and check sheet names
        wb = load_workbook(dest_path)
        assert len(wb.sheetnames) == 5
        assert wb.sheetnames == ["Summary", "Valuation", "Health", "Growth", "Red Flags"]

        # Verify Summary sheet contains company name
        ws_summary = wb["Summary"]
        found_company = False
        for row in ws_summary.iter_rows(values_only=True):
            if "Apple Inc." in row:
                found_company = True
                break
        assert found_company


def test_excel_export_graceful_on_missing_data() -> None:
    # Construct a ScanResult with None fields
    result = ScanResult(
        ticker="XYZ",
        company="XYZ Corp",
        exchange="NYSE",
        scan_duration_ms=0,
        valuation=ValuationResult(
            pe_ratio=None,
            pb_ratio=None,
            ev_ebitda=None,
            vs_sector=None,
            flag=None
        ),
        health=HealthResult(
            piotroski_f=None,
            altman_z=None,
            debt_to_equity=None,
            current_ratio=None,
            interest_coverage=None,
            flag=None
        ),
        growth=GrowthResult(
            revenue_cagr_3y=None,
            pat_cagr_3y=None,
            margin_trend=None,
            fcf_yield=None,
            flag=None
        ),
        red_flags=[],
        highest_concern="low",
        analyst_brief="No details."
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        dest_path = os.path.join(tmpdir, "test_empty_report.xlsx")

        # Export should not crash
        exporter = ExcelExporter()
        exporter.export(result, dest_path)

        assert os.path.exists(dest_path)

        wb = load_workbook(dest_path)
        assert len(wb.sheetnames) == 5

        # Verify empty red flags cell has the correct message
        ws_flags = wb["Red Flags"]
        assert ws_flags["A2"].value == "No red flags detected"
